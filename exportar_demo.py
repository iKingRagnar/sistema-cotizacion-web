# Genera seed-demo.json desde SistemaGestion_Demo.xlsm (clientes, refacciones, máquinas, incidentes, bitácoras)
import openpyxl
import json
import os
import re

xlsm = r'c:\Users\ragna\Downloads\SistemaGestion_Demo.xlsm'
if not os.path.isfile(xlsm):
    xlsm = os.path.join(os.path.dirname(__file__), '..', 'SistemaGestion_Demo.xlsm')
wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)

def clean(v):
    if v is None: return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return int(v) if float(v).is_integer() else round(float(v), 2)
    s = str(v).strip()
    return s if s else None

def fecha_solo(val):
    if not val: return None
    s = str(val)
    if ' ' in s: s = s.split(' ')[0]
    if re.match(r'\d{4}-\d{2}-\d{2}', s): return s
    return s[:10] if len(s) >= 10 else s

# --- CLIENTES ---
def get_clientes():
    ws = wb['CAT_CLIENTES']
    rows = list(ws.iter_rows(values_only=True))[1:]
    out = []
    for row in rows:
        if not row or row[0] is None: continue
        out.append({
            'codigo': 'CLI-%s' % int(row[0]),
            'nombre': clean(row[1]),
            'rfc': clean(row[2]),
            'contacto': clean(row[3]),
            'telefono': clean(row[4]),
            'email': clean(row[5]),
            'direccion': clean(row[6]),
            'ciudad': clean(row[7]) if len(row) > 7 else None,
        })
    return [c for c in out if c.get('nombre')]

# --- REFACCIONES ---
def get_refacciones():
    ws = wb['CAT_REFACCIONES']
    rows = list(ws.iter_rows(values_only=True))[1:]
    out = []
    for row in rows:
        if not row or row[0] is None: continue
        out.append({
            'codigo': clean(row[1]) or ('REF-%s' % row[0]),
            'descripcion': clean(row[2]) or '-',
            'marca': clean(row[3]),
            'origen': clean(row[4]),
            'precio_unitario': clean(row[5]) if len(row) > 5 and row[5] is not None else 0,
            'unidad': clean(row[6]) if len(row) > 6 else 'PZA',
        })
    return [r for r in out if r.get('codigo')]

# --- MÁQUINAS ---
def get_maquinas():
    ws = wb['CAT_MAQUINAS']
    rows = list(ws.iter_rows(values_only=True))[1:]
    out = []
    for row in rows:
        if not row or row[0] is None: continue
        out.append({
            'nombre': clean(row[1]),
            'marca': clean(row[2]),
            'modelo': clean(row[3]),
            'numero_serie': clean(row[4]) if len(row) > 4 else None,
            'cliente_id': int(row[5]) if len(row) > 5 and row[5] is not None else None,
            'ubicacion': clean(row[6]) if len(row) > 6 else None,
        })
    return [m for m in out if m.get('nombre')]

# --- INCIDENTES (FOLIO, FECHA, CLIENTE, MÁQUINA, DESCRIPCIÓN, PRIORIDAD, TÉCNICO, ESTATUS, ...)
def get_incidentes():
    if 'INCIDENTES' not in wb.sheetnames: return []
    ws = wb['INCIDENTES']
    rows = list(ws.iter_rows(values_only=True))[1:]
    out = []
    for row in rows:
        if not row or not clean(row[0]): continue
        prio = (clean(row[5]) or '').upper()
        if 'P1' in prio or 'CRITIC' in prio: prioridad = 'critica'
        elif 'P2' in prio or 'ALTA' in prio: prioridad = 'alta'
        elif 'P3' in prio or 'MEDIA' in prio: prioridad = 'media'
        else: prioridad = 'baja'
        est = (clean(row[7]) or 'Abierto').lower()
        if 'resuelto' in est or 'cerrado' in est: estatus = 'cerrado'
        elif 'proceso' in est: estatus = 'en_proceso'
        elif 'cancelado' in est: estatus = 'cancelado'
        else: estatus = 'abierto'
        out.append({
            'folio': clean(row[0]),
            'fecha_reporte': fecha_solo(row[1]),
            'cliente_nombre': clean(row[2]),
            'maquina_nombre': clean(row[3]),
            'descripcion': clean(row[4]) or '-',
            'prioridad': prioridad,
            'tecnico_responsable': clean(row[6]),
            'estatus': estatus,
        })
    return out

# --- BITÁCORAS (FOLIO_BITACORA, FOLIO_INC_REF, FECHA, TÉCNICO, ..., ACTIVIDADES, REFACCIONES USADAS, HORAS, ...)
def get_bitacoras():
    if 'BITACORAS' not in wb.sheetnames: return []
    ws = wb['BITACORAS']
    rows = list(ws.iter_rows(values_only=True))[1:]
    out = []
    for row in rows:
        if not row or not clean(row[1]): continue  # folio_inc_ref
        try:
            horas = float(row[8]) if len(row) > 8 and row[8] is not None else 0
        except (TypeError, ValueError):
            horas = 0
        out.append({
            'folio_incidente': clean(row[1]),  # FOLIO_INC_REF -> para vincular a incidente
            'fecha': fecha_solo(row[2]),
            'tecnico': clean(row[3]),
            'actividades': clean(row[6]) if len(row) > 6 else None,
            'materiales_usados': clean(row[7]) if len(row) > 7 else None,
            'tiempo_horas': horas,
        })
    return out

clientes = get_clientes()
refacciones = get_refacciones()
maquinas = get_maquinas()
incidentes = get_incidentes()
bitacoras = get_bitacoras()
wb.close()

seed = {
    'clientes': clientes,
    'refacciones': refacciones,
    'maquinas': maquinas,
    'incidentes': incidentes,
    'bitacoras': bitacoras,
}
out_path = os.path.join(os.path.dirname(__file__), 'seed-demo.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(seed, f, ensure_ascii=False, indent=2)
print('Generado:', out_path)
print('Clientes:', len(clientes), 'Refacciones:', len(refacciones), 'Maquinas:', len(maquinas))
print('Incidentes:', len(incidentes), 'Bitacoras:', len(bitacoras))
