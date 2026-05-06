# Inspeccionar hojas COTIZACION, INCIDENTES, BITACORAS
import openpyxl
import json
import os
xlsm = r'c:\Users\ragna\Downloads\SistemaGestion_Demo.xlsm'
if not os.path.isfile(xlsm):
    xlsm = os.path.join(os.path.dirname(__file__), '..', 'SistemaGestion_Demo.xlsm')
wb = openpyxl.load_workbook(xlsm, read_only=True, data_only=True)
out = []
for name in ['COTIZACION', 'INCIDENTES', 'BITACORAS']:
    if name not in wb.sheetnames:
        out.append(name + ': NO EXISTE')
        continue
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    # Guardar como valores "seguros" para JSON
    header = [str(c) if c is not None else '' for c in rows[0]] if rows else []
    sample = []
    for row in rows[1:6]:
        sample.append([str(v)[:50] if v is not None else None for v in row])
    out.append({'sheet': name, 'header': header, 'sample': sample, 'total_rows': len(rows)-1})
wb.close()
with open(os.path.join(os.path.dirname(__file__), 'inspect_result.json'), 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print('OK inspect_result.json')
